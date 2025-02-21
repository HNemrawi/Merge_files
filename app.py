import streamlit as st
import openpyxl
import shutil
import os
import tempfile

def sum_excel_into_destination(files_folder, temp_path, destination_path):
    # Create a new destination file as a copy of temp.xlsx
    shutil.copy(temp_path, destination_path)
    wb_dest = openpyxl.load_workbook(destination_path)
    
    # Get all Excel files from the specified folder
    source_files = [os.path.join(files_folder, f) for f in os.listdir(files_folder) if f.endswith(".xlsx")]
    
    # Load all source workbooks
    workbooks = [openpyxl.load_workbook(file) for file in source_files]
    
    # Iterate through each sheet in the destination workbook
    for sheet_name in wb_dest.sheetnames:
        # Ensure this sheet exists in all source workbooks
        if all(sheet_name in wb.sheetnames for wb in workbooks):
            ws_dest = wb_dest[sheet_name]
            source_sheets = [wb[sheet_name] for wb in workbooks]
            
            # Determine the max row and column size from the destination sheet
            max_row = ws_dest.max_row
            max_col = ws_dest.max_column
            
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    # Sum up all numeric values from source files
                    total = 0
                    has_numeric = False
                    for ws in source_sheets:
                        val = ws.cell(row=row, column=col).value
                        if isinstance(val, (int, float)):
                            total += val
                            has_numeric = True
                    # If at least one numeric value was found, update the destination cell
                    if has_numeric:
                        ws_dest.cell(row=row, column=col).value = total

    # Save the updated destination workbook
    wb_dest.save(destination_path)
    return destination_path

st.title("Excel Summation App")

st.write("Upload as many Excel files (.xlsx) as you need to sum them up based on the template.")

# Allow multiple Excel files to be uploaded
uploaded_files = st.file_uploader("Upload Excel files", type=["xlsx"], accept_multiple_files=True)

if uploaded_files:
    # Use a temporary directory to save uploaded files
    with tempfile.TemporaryDirectory() as upload_dir:
        for uploaded_file in uploaded_files:
            file_path = os.path.join(upload_dir, uploaded_file.name)
            with open(file_path, "wb") as f:
                f.write(uploaded_file.getbuffer())
        
        # Define the path for the destination file in the temporary directory
        destination_path = os.path.join(upload_dir, "destination.xlsx")
        # temp.xlsx should be in the same directory as the Streamlit app
        temp_template_path = "temp.xlsx"
        
        if st.button("Process Files"):
            try:
                result_path = sum_excel_into_destination(upload_dir, temp_template_path, destination_path)
                st.success("Summation completed!")
                # Read the resulting file in binary mode to enable download
                with open(result_path, "rb") as f:
                    st.download_button(
                        label="Download Result",
                        data=f,
                        file_name="destination.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e:
                st.error(f"An error occurred: {e}")